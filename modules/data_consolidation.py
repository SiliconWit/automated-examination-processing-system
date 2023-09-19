import os, toml, re, sys, json
import pandas as pd
from .file_processing import *
from .utilities import *
from collections import Counter
from .rule_engine import *

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from reportlab.pdfgen import canvas

import inflect

import PyPDF2 # For combining PDFs



config_path = "config.toml"  # Specify the path to your TOML configuration file
# Load the configuration from the TOML file
config = toml.load(config_path)
input_folder_path = config["input_folder"]["path"]

def consolidate_mark_sheet(mechatronics_units_path, input_folder_path, consolidated_excel_output_path, pass_list_pdf_output_path, supp_list_pdf_output_path, senate_doc_pdf_output_path, config_path):
    # Get a list of all Excel files in the input folder

    center_names = fetch_center_names(input_folder_path)
    log_print("Exam files: {}".format(center_names))





    mechatronics_json_data = json.load(open(mechatronics_units_path))

    # Step 1: Check if all files are .xlsx
    check_xlsx_files(input_folder_path)

    # Step 2: Check if filenames match Unit Codes
    # unit_codes = {unit["Unit Code"] for year_data in mechatronics_json_data.values() for semesters in year_data.values() if isinstance(semesters, (list, dict)) for unit in semesters if isinstance(unit, dict)}
    unit_codes = center_names
    check_filenames_match_units(input_folder_path, unit_codes)

    # Step 3: Check if unit codes belong to a single year
    year = check_unit_codes_single_year(unit_codes, mechatronics_json_data)

    log_print(f"All filenames match Unit Codes and belong to the year: {year}")
    yr_to_process = year 





    # Get the column order from the configuration
    desired_columns = config["column_order"]["columns"]
    additional_columns = config["additional_columns"]["columns"]

    # Generate the center names dynamically
    center_names = fetch_center_names(config["input_folder"]["path"])
    center_columns = list(center_names)
    # print(center_columns)

    # Combine all columns: desired columns, dynamic center columns, additional columns
    new_column_order = desired_columns + center_columns + additional_columns

    # Create an empty DataFrame using the generated column list
    consolidated_df = pd.DataFrame(columns=new_column_order)

    # Loop through each Excel file and consolidate the data
    course_code = loop_to_consolidate(excel_files, consolidated_df, collected_data)


    # Group collected data by Reg. No.
    grouped_data = {}
    for course, file_course_code, reg_no, name, internal_marks in collected_data:
        # print(collected_data)
        if reg_no in grouped_data:
            grouped_data[reg_no].append((name, file_course_code, internal_marks))
        else:
            grouped_data[reg_no] = [(name, file_course_code, internal_marks)]

    # open('../collected_data2.txt', 'w').writelines('\n'.join(map(str, collected_data)) + '\n')
    # open('../grouped_data.txt', 'w').writelines('\n'.join(map(str, grouped_data.items())) + '\n')
    # print(course_code)
    # Consolidate names for the same Reg. No.
    consolidated_names = {}
    # print(grouped_data.items())



    for student_id, name_code_mark in grouped_data.items():
        # print(grouped_data.items())
        # print(name_code_mark)
        consolidated_names[student_id] = {'Name': '', 'Code': '', 'Marks': []}
        
        # name_counts = Counter([name for name, code, mark in name_code_mark if name and not pd.isna(name) and name.strip() != ""])
        # if name_counts:
        #     most_common_name = name_counts.most_common(1)[0][0]
        #     consolidated_names[student_id]['Name'] = most_common_name

        # Advanced name format 
        name_counts = Counter([name for name, code, mark in name_code_mark if name and not pd.isna(name) and name.strip() != ""])
        if name_counts:
            most_common_name = name_counts.most_common(1)[0][0]
            # Split the most common name into parts (first and last name)
            name_parts = most_common_name.split()
            if len(name_parts) >= 2:
                # Join the parts in title case except for the last name in all caps
                formatted_name = ' '.join([part.capitalize() if i < len(name_parts) - 1 else part.upper() for i, part in enumerate(name_parts)])
                consolidated_names[student_id]['Name'] = formatted_name

        course_code_name = [code for name, code, mark in name_code_mark ]
        if course_code_name:
            consolidated_names[student_id]['Code'] = course_code_name
        
        marks = [mark for name, code, mark in name_code_mark ]
        if marks:
            consolidated_names[student_id]['Marks'] += marks

    # open('../consolidated_names.txt', 'w').writelines('\n'.join(map(str, consolidated_names.items())) + '\n')

    # Create an empty DataFrame
    consolidated_data_df = pd.DataFrame(columns=['Reg. No.', 'Name'] + course_code)

    # Populate the DataFrame
    for reg_no, data in consolidated_names.items():
        name = data['Name']
        codes = data.get('Code', [])
        marks = data.get('Marks', [])
        
        # Create a dictionary to hold marks for each course code
        code_marks = {code: mark for code, mark in zip(codes, marks)}
        
        # Fill missing exam codes with empty strings and add to the row
        row = [reg_no, name] + [code_marks.get(code, '') for code in course_code]
        
        consolidated_data_df.loc[len(consolidated_data_df)] = row


    # Sort the consolidated data based on 'Reg. No.'
    consolidated_data_df['Sort Key'] = consolidated_data_df['Reg. No.'].apply(sort_key)
    consolidated_data_df = consolidated_data_df.sort_values(by='Sort Key')

    # Drop the temporary 'Sort Key' column
    consolidated_data_df.drop(columns=['Sort Key'], inplace=True)


    # # Get units done in semester_4_1 and yr4_2nd_semester
    # semester_4_1_units = config['semester_order']['semester_4_1']
    # semester_4_2_units = config['semester_order']['semester_4_2']

    # Extract unit codes from the 4th Year 1st Semester and 4th Year 2nd Semester
    x_year_1st_semester = mechatronics_json_data.get(yr_to_process, {}).get("1st Semester", [])
    x_year_2nd_semester = mechatronics_json_data.get(yr_to_process, {}).get("2nd Semester", [])

    # Extract unit codes from the units
    x_yr_1st_semester_codes = [unit["Unit Code"] for unit in x_year_1st_semester]
    x_yr_2nd_semester_codes = [unit["Unit Code"] for unit in x_year_2nd_semester]
    #print(x_yr_1st_semester_codes)
    # print(least_common_unit_codes)
    print(x_yr_2nd_semester_codes)

    # Check if 'least_common_unit_codes' contains non-empty elements before inserting into my_list
    if any(elem.strip() for elem in least_common_unit_codes):
        x_yr_1st_semester_codes.insert(0, *least_common_unit_codes) # Insert before 

    # Initialize a list to store the rearranged course codes
    rearranged_course_code = []

    # Rearrange course codes based on semester order
    for unit in x_yr_1st_semester_codes:
        if unit in course_code:
            rearranged_course_code.append(unit)
        else:
            matched_unit_name = find_unit_name(mechatronics_units_path, unit)
            log_print(f"Unit {unit} {matched_unit_name} was NOT done in {yr_to_process} 1st semester")

    for unit in x_yr_2nd_semester_codes:
        if unit in course_code:
            rearranged_course_code.append(unit)
        else:
            matched_unit_name = find_unit_name(mechatronics_units_path, unit)
            log_print(f"Unit {unit} {matched_unit_name}  was NOT done in {yr_to_process} 2nd semester")

    # Define the columns to consider for checking for missing marks
    columns_to_check = rearranged_course_code

    # Replace empty strings ('' or ' ') with nan in the specified columns
    consolidated_data_df[columns_to_check] = consolidated_data_df[columns_to_check].replace(['', ' '], np.nan)

    # Use dropna to remove rows with missing values in specified columns
    consolidated_data_df = consolidated_data_df.dropna(subset=columns_to_check, how='all')

    # Reset the index after removing rows
    consolidated_data_df = consolidated_data_df.reset_index(drop=True)

    # Add 'Ser. No.' column with a count
    consolidated_data_df['Ser. No.'] = range(1, len(consolidated_data_df) + 1)

    # Combine all columns: desired columns, dynamic center columns, additional columns
    new_column_order = desired_columns + rearranged_course_code + additional_columns


    # Reorder the columns in the DataFrame using the reindex method
    consolidated_data_df = consolidated_data_df.reindex(columns=new_column_order)



    # Calculate TU (Total Units), Total, and Mean for each row
    for index, row in consolidated_data_df.iterrows():
        total_units = 0
        total_marks = 0
        
        # Calculate total units and total marks for the rearranged_course_code columns
        for code in rearranged_course_code:
            if not np.isnan(row[code]):
                total_units += 1
                total_marks += row[code]
        
        # Fill in TU and Total columns
        consolidated_data_df.at[index, 'TU'] = total_units
        consolidated_data_df.at[index, 'Total'] = total_marks
        
        # Calculate and fill Mean column if all units were done
        if total_units == len(rearranged_course_code):
            mean = total_marks / total_units
            consolidated_data_df.at[index, 'Mean'] = "{:.2f}".format(mean)


    calculate_grade(mean)

    # Add Grade column and fill it based on the Mean column
    consolidated_data_df['Grade'] = consolidated_data_df['Mean'].apply(calculate_grade)


    # Define a function to calculate the status and count supplementaries and blank cases
    def calculate_recommendation(row):
        supplementaries = [1 for code in rearranged_course_code if
                        (isinstance(row[code], float) and row[code] < 40) ]

        blank_cases = [1 for code in rearranged_course_code if
                        isinstance(row[code], (str, float, np.nan)) and (row[code] == '' or pd.isna(row[code]) or (
                                isinstance(row[code], str) and row[code].isspace()))]

        status = []

        if supplementaries:
            status.append(f'SUPP = {sum(supplementaries)} UNIT{"S" if sum(supplementaries) > 1 else ""}')
        if blank_cases:
            status.append(f'BLANK = {sum(blank_cases)} UNIT{"S" if sum(blank_cases) > 1 else ""}')

        return ', '.join(status) if status else 'PASS'


    # Add Status column and fill it based on the rearranged_course_code columns
    consolidated_data_df['Status'] = consolidated_data_df.apply(calculate_recommendation, axis=1)






    # Filter the DataFrame to include only students who passed
    passed_students_df = consolidated_data_df[consolidated_data_df['Status'] == 'PASS']
    # print(passed_students_df)

    # Custom function to check if the Status contains 'SUPP' and does not contain 'BLANK' and units information
    def has_supp_units(status):
        return 'SUPP' in status and 'BLANK' not in status and any(word.isdigit() for word in status.split())

    # Filter the DataFrame using the custom function
    supp_students_df = consolidated_data_df[consolidated_data_df['Status'].apply(has_supp_units)]


    # Filter the DataFrame to include only students who passed
    blank_students_df = consolidated_data_df[consolidated_data_df['Status'] == 'BLANK']


    # Select the 'Ser. No.', 'Reg. No.' and 'Name' columns
    passed_students_list = passed_students_df[['Reg. No.', 'Name']]
    supp_students_list = supp_students_df[['Reg. No.', 'Name']]

    # Reset the index after removing rows
    passed_students_list = passed_students_list.reset_index(drop=True)
    supp_students_list = supp_students_list.reset_index(drop=True)

    # Add 'Ser. No.' column with a count
    passed_students_list['Ser. No.'] = range(1, len(passed_students_list) + 1)
    supp_students_list['Ser. No.'] = range(1, len(supp_students_list) + 1)


    # Combine all columns: desired columns 
    pass_columns_order = desired_columns 

    # Reorder the columns in the DataFrame using the reindex method
    passed_students_list = passed_students_list.reindex(columns=pass_columns_order)
    supp_students_list = supp_students_list.reindex(columns=desired_columns)
    # print(len(passed_students_list['Ser. No.']))


    # Select the 'Ser. No.', 'Reg. No.' and 'Name' columns
    # supp_students_list = supp_students_df[['Ser. No.', 'Reg. No.', 'Name']]

    # Select the 'Ser. No.', 'Reg. No.' and 'Name' columns
    blank_students_list = blank_students_df[['Ser. No.', 'Reg. No.', 'Name']]

    # Save the filtered data to a new .csv file
    # supp_students_list.to_csv('../supp_students.csv', index=False)





    # Create an inflect engine
    p = inflect.engine()

    # # Register the Palatino font
    # pdfmetrics.registerFont(TTFont('Palatino', 'fonts/palatino-regular.ttf'))
    # addMapping('Palatino', 0, 0, 'Palatino')  # Map the font name


    # Create a PDF documents
    # supp_list_filename = 'supp_students.pdf'
    # blank_list_filename = 'blank_students.pdf'

    # Get the base name (name of the file without extension)
    pass_list_pdf_name = os.path.splitext(os.path.basename(pass_list_pdf_output_path))[0]
    supp_list_pdf_name = os.path.splitext(os.path.basename(supp_list_pdf_output_path))[0]

    # Define the pass_content for the PDF
    pass_content = []
    supp_content = []

    # Get the template from the config
    doc_title = config["document_title"]["document_title"]
    university_name = config["senate_documents_details"]["university_name"]
    school_of = config["senate_documents_details"]["school_of"]
    department_of = config["senate_documents_details"]["department_of"]
    course_name = config["senate_documents_details"]["course_name"]
    academic_year = config["senate_documents_details"]["academic_year"]

    # year_of_study = config["senate_documents_details"]["year_of_study"]
    year_of_study = yr_to_process

    # Extract the numeric part using the first character of year_of_study
    # year_of_study_int will contain the integer value (1, 2, 3, 4, or 5)
    year_of_study_int = int(year_of_study[0])
    # print(year_of_study_int)

    year_of_study = p.number_to_words(p.ordinal(year_of_study_int)).upper()
    year_of_study_plus = p.number_to_words(p.ordinal(int(year_of_study_int)+1)).capitalize() 
    # semester_of_study = config["senate_documents_details"]["semester_of_study"]
    
    # print(center_names)
    # print(x_yr_2nd_semester_codes)
    x_yr_2nd_semester_codes_set = set(x_yr_2nd_semester_codes)
    # Calculate the number of units in 'center_names' not available in x_yr_2nd_semester_codes_set
    unavailable_units = len(center_names - x_yr_2nd_semester_codes_set)

    # Define a threshold for "most" (you can adjust this as needed)
    semester_threshold = len(center_names) // 2  # Half (50%) of the units in center_names

    # Check if most units in center_names are not available in x_yr_2nd_semester_codes_set
    if unavailable_units > semester_threshold:
        semester_of_study = "1"
        semester_of_study_plus = int(semester_of_study[0])+1
        semester_of_study = p.number_to_words(p.ordinal(semester_of_study)).upper()
        pass_intro_recommendation_txt = "<b>they proceed to the {} Semester</b>".format(p.number_to_words(p.ordinal(semester_of_study_plus)).capitalize()) 
        log_print("This seems to be the 1st Semester.")
    else:
        semester_of_study = "2"
        semester_of_study = p.number_to_words(p.ordinal(semester_of_study)).upper()
        pass_intro_recommendation_txt = "<b>they proceed to the {} Year of study</b>".format(year_of_study_plus)
        log_print("This seems to be 2nd Semester.")
        log_print("All or most of 2nd Semester units were done.")
    
    
    doc_title = doc_title.format(university_name, school_of, department_of, course_name, academic_year, year_of_study, semester_of_study)
    
    pass_list_intro = config["pass_list_introduction"]["pass_list_intro_content"]
    supp_list_intro = config["supp_list_introduction"]["supp_list_intro_content"]

    doc_sign_text = config["document_signature_text"]["document_signature_content"]
    
    # The number of candidates
    pass_num_candidates = len(passed_students_list['Ser. No.'])  
    supp_num_candidates = len(supp_students_list['Ser. No.'])  

    # Convert the numeric value into words (e.g., 50 to "Fifty")
    pass_num_words = p.number_to_words(pass_num_candidates).capitalize()
    supp_num_words = p.number_to_words(supp_num_candidates).capitalize()

    # Fill in the template with the actual value
    pass_list_intro_text = pass_list_intro.format(pass_num_words,pass_num_candidates, school_of, academic_year, year_of_study.capitalize(), semester_of_study.capitalize(), course_name, school_of, pass_intro_recommendation_txt)
    supp_list_intro_text = supp_list_intro.format(supp_num_words, supp_num_candidates, year_of_study.capitalize(), semester_of_study.capitalize() , course_name , academic_year, school_of)

    # Add a letterhead as a Paragraph
    styles = getSampleStyleSheet()
    # letterhead_text = "Department of XYZ University\nList of {} Passed Students".format(len(passed_students_list['Ser. No.']))
    # doc_title_text = Paragraph(doc_title, styles['Title'])

    # Define a custom style for the title
    title_style = ParagraphStyle(
        name='TitleStyle',
        fontName='Helvetica-Bold',  # Use Times-Bold for bold
        fontSize=12,  # Adjust the font size as needed
        alignment=1,  # Center alignment (1 for center)
    )

    # Create the Paragraph using the custom title style
    doc_title_text = Paragraph(doc_title, title_style)

    # Append the Paragraph to the pass_content
    pass_content.append(doc_title_text)
    supp_content.append(doc_title_text)

    pass_list_introduction = Paragraph(pass_list_intro_text, styles['Normal'])
    supp_list_introduction = Paragraph(supp_list_intro_text, styles['Normal'])

    # Add a spacer
    pass_content.append(Spacer(1, 12))
    supp_content.append(Spacer(1, 12))

    pass_content.append(Paragraph("<u>PASS LIST</u>", title_style))
    supp_content.append(Paragraph("<u>SUPPLEMENTARY LIST</u>", title_style))

    # Add a spacer
    # pass_content.append(Spacer(1, 12))
    # supp_content.append(Spacer(1, 12))

    pass_content.append(pass_list_introduction)
    supp_content.append(supp_list_introduction)

    # Add a spacer
    pass_content.append(Spacer(1, 12))
    supp_content.append(Spacer(1, 12))

    # Create a table for the passed students
    pass_data = [passed_students_list.columns.tolist()] + passed_students_list.values.tolist()
    supp_data = [supp_students_list.columns.tolist()] + supp_students_list.values.tolist()
    
    pass_table = Table(pass_data)
    supp_table = Table(supp_data)

    # Add style to the pass_table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        # Left-align the "Name" column (index 2)
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
    ])

    pass_table.setStyle(style)
    supp_table.setStyle(style)

    pass_content.append(pass_table)
    supp_content.append(supp_table)

    # Add a spacer
    pass_content.append(Spacer(1, 12))
    supp_content.append(Spacer(1, 12))

    doc_sign_txt = Paragraph(doc_sign_text, styles['Normal'])
    pass_content.append(doc_sign_txt)
    supp_content.append(doc_sign_txt)



    def generate_pdf_with_centered_page_numbers(pdf_output_path, title, pass_content):
        doc = SimpleDocTemplate(pdf_output_path, pagesize=letter, bottomMargin=50)
        # Create a SimpleDocTemplate with specified metadata
        doc.title = title
        doc.subject = "Automatic Exams Processing System (AEPS) Results"
        doc.author = "SiliconWit"
        doc.creator = "SiliconWit Open AEPS"
        doc.producer = "https://siliconwit.com/"
        doc.keywords = "Exams Processing"

        def add_centered_page_numbers(canvas, doc):
            page_num = canvas.getPageNumber()
            page_text = f"Page {page_num}"
            canvas.setFont("Times-Roman", 9)  # or set the font to Palatino or inbuild Courier
            canvas.drawCentredString(letter[0] / 2, 30, page_text)  # Center the page numbers at the bottom

        # Create the custom canvas with centered page numbers
        c = canvas.Canvas(pdf_output_path, pagesize=letter)
        c.showPage()
        c.save()

        # Add your pass_content to the PDF
        doc.build( pass_content, onFirstPage=add_centered_page_numbers, onLaterPages=add_centered_page_numbers)
        log_print(f"PDF report saved as '{pdf_output_path}'")


    generate_pdf_with_centered_page_numbers(pass_list_pdf_output_path, pass_list_pdf_name, pass_content)
    generate_pdf_with_centered_page_numbers(supp_list_pdf_output_path, supp_list_pdf_name, supp_content)

    # Combine PDFs 
    # List of PDF files to combine
    input_pdfs = [pass_list_pdf_output_path, supp_list_pdf_output_path]

    # Output file name for the combined PDF
    # Update senate doc name
    general_senate_pdf_path = senate_doc_pdf_output_path
    yr_processed = yr_to_process

    # Remove spaces from yr_processed and replace them with underscores
    yr_processed = yr_processed.replace(" ", "_")

    # Convert the characters in yr_processed to lowercase
    yr_processed = yr_processed.lower()+"_"
    # print(semester_of_study)
    to_numeric_conversions = {"FIRST": "1st", "SECOND": "2nd", "THIRD": "3rd", "FOURTH": "4th"}

    sem_processed = to_numeric_conversions.get(semester_of_study, semester_of_study)+"_sem"
    # print(sem_processed)

    # Split the file extension from general_senate_pdf_path
    base_name, extension = os.path.splitext(general_senate_pdf_path)

    # Combine base_name, yr_processed, and the original extension to create output_pdf_file
    output_pdf_file = f"{base_name}_{yr_processed}{sem_processed}{extension}"



    # Create a PDF merger object
    pdf_merger = PyPDF2.PdfFileMerger()

    # Append each input PDF to the merger
    for pdf_file in input_pdfs:
        pdf_merger.append(pdf_file)

    # Write the merged PDF to the output file
    with open(output_pdf_file, 'wb') as output_pdf:
        pdf_merger.write(output_pdf)



    # Create a new Excel file and save the consolidated data
    wb = Workbook()
    ws = wb.active

    # Define fill colors 
    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


    # Add data to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(consolidated_data_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Check if the cell contains a pass status or blank case and apply the red fill
            if isinstance(value, str) and ('PASS' in value or value == 'PASS'):
                cell.fill = light_green_fill

            # Check for marks below 40 and color them in a more intense red
            if isinstance(value, (int, float)) and value < 40 and ws.cell(row=1, column=c_idx).value in columns_to_check:
                cell.fill = light_blue_fill

            # Check if the cell contains a supplementary status or blank case and apply the red fill
            if isinstance(value, str) and ('SUPP' in value or value == 'SUPP'):
                cell.fill = light_blue_fill

            # Check if the cell contains a blank status or blank case and apply the red fill
            if isinstance(value, str) and ('BLANK' in value or value == 'BLANK'):
                cell.fill = red_fill

            # Check for empty strings ('' or ' '), spaces, or nan and color them grey
            elif isinstance(value, str) and (value == '' or value.isspace()):
                cell.fill = red_fill
            elif isinstance(value, float) and np.isnan(value):
                cell.fill = red_fill

    # Adjust column widths to fit the data
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook
    wb.save(consolidated_excel_output_path)



    log_print(f"Consolidated mark sheet saved as '{consolidated_excel_output_path}'.")

    # Print files without "REG. NO." cell
    if files_without_reg_no:
        log_print("Files without 'REG. NO.' cell:")
        for file in files_without_reg_no:
            log_print(file)
